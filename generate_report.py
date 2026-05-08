import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
sub_header_font = Font(name='Microsoft YaHei', bold=True, size=10, color='2F5496')
data_font = Font(name='Microsoft YaHei', size=10)
check_font = Font(name='Microsoft YaHei', size=10, color='00B050')
cross_font = Font(name='Microsoft YaHei', size=10, color='FF0000')
title_font = Font(name='Microsoft YaHei', bold=True, size=14, color='2F5496')
section_font = Font(name='Microsoft YaHei', bold=True, size=11, color='2F5496')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_sub_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = center_align
        cell.border = thin_border

def write_row(ws, row, data, bold_first=False):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        is_check = isinstance(val, str) and val.startswith('✅')
        is_cross = isinstance(val, str) and val.startswith('❌')
        cell.font = check_font if is_check else (cross_font if is_cross else data_font)
        cell.alignment = center_align if col > 1 else left_align
        cell.border = thin_border
    if bold_first:
        ws.cell(row=row, column=1).font = Font(name='Microsoft YaHei', bold=True, size=10)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 1: 总览对比表
# ============================================================
ws1 = wb.active
ws1.title = "总览对比表"
ws1.merge_cells('A1:H1')
ws1['A1'] = '海外网红营销平台竞品调研报告（官网验证版）'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

headers = ['平台', '数据库规模', '支持平台', '起步价(月)', '内置邮件外联', 'CRM', '电商集成', '假粉检测']
for col, h in enumerate(headers, 1):
    ws1.cell(row=3, column=col, value=h)
style_header(ws1, 3, 8)

data = [
    ['NoxInfluencer', '1亿+（35万+认证）', 'YouTube/TikTok/IG/Twitter', '-', '✅ 邮件+WhatsApp', '✅', '✅ Shopify', '-'],
    ['Modash', '3.5亿-3.8亿+', 'IG/TikTok/YouTube', '$199(年付)/$299(月付)', '✅ Gmail/Outlook同步+批量+Drip', '✅ 基础', '✅ Shopify+付款+联盟', '✅ AQS+假粉率'],
    ['Heepsy', '1100万+', 'IG/TikTok/YouTube', '$69', '✅ 批量联系+消息+历史', '✅', '✅ Shopify/Amazon/TikTok Shop', '✅ 真实性评分+品牌安全'],
    ['Grin', '75万+', '仅确认IG', '$399', '✅ Gmail/Klaviyo+模板序列', '✅', '✅ Shopify(为主)', '-'],
    ['CreatorIQ', '2000万+', '-', '不公开', '✅ Recruit模块', '✅', '-', '✅ SafeIQ品牌安全'],
    ['Aspire', '数百万', '-', '不公开', '✅ Contact Hub+工作流', '✅', '✅ Shopify(深度)+WooCommerce', '-'],
    ['Traackr', '数百万/数十亿内容', '-', '不公开', '未在官网展示', '-', '-', '-'],
    ['Meltwater(原Klear)', '3000万+', 'IG/TikTok/Twitter/FB/YT/Pinterest/博客', '不公开', '✅', '✅', '✅ Shopify+WooCommerce+追踪像素', '✅ 机器人检测'],
    ['Influencity', '2亿+（TikTok 1.3亿+）', 'TikTok/IG/YouTube/Twitch', '不公开', '✅ IRM模块', '✅', '✅ 电商店铺集成', '-'],
    ['SocialBook', '3亿+', 'YouTube/IG/TikTok', '-', '-', '-', '-', '✅ IG假粉检测'],
    ['FastMoss', '2.5亿达人/5亿商品', 'TikTok', '-', '✅ MossCreator批量邮件', '-', 'TikTok Shop', '-'],
    ['EchoTik', '1亿达人/1.8亿商品', 'TikTok', '$9.9(Basic)', '✅ 开发信生成(AI)', '-', 'TikTok Shop', '-'],
    ['Kalodata', '2.5亿达人/2亿商品/4亿视频', 'TikTok Shop', '-', '-', '-', 'TikTok Shop', '-'],
    ['Social Blade', '1.6亿创作者', 'YouTube/TikTok/IG/Twitch/FB', '-', '❌ 分析工具非营销平台', '❌', '❌', '❌'],
]
for i, row_data in enumerate(data):
    write_row(ws1, i + 4, row_data, bold_first=True)
set_col_widths(ws1, [22, 28, 40, 25, 35, 10, 30, 22])

# ============================================================
# Sheet 2: 网红发现 - 平台覆盖与数据库
# ============================================================
ws2 = wb.create_sheet("网红发现-平台与数据库")
ws2.merge_cells('A1:K1')
ws2['A1'] = '网红发现能力对比 — 平台覆盖与数据库规模'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')
ws2.row_dimensions[1].height = 30

headers2 = ['平台', '数据库规模', 'Instagram', 'YouTube', 'TikTok', 'Twitter/X', 'Facebook', 'Pinterest', 'LinkedIn', 'Twitch', '博客']
for col, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=col, value=h)
style_header(ws2, 3, 11)

platform_data = [
    ['NoxInfluencer', '1亿+（35万+认证）', '✅', '✅(核心)', '✅', '✅', '-', '-', '-', '-', '-'],
    ['Modash', '3.5亿-3.8亿+', '✅', '✅', '✅', '-', '-', '-', '-', '-', '-'],
    ['Heepsy', '1100万+', '✅(核心)', '✅', '✅', '-', '-', '-', '-', '-', '-'],
    ['Grin', '75万+', '✅(确认)', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['CreatorIQ', '2000万+', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['Aspire', '数百万(激活30万)', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['Traackr', '数百万/数十亿内容', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['Meltwater(原Klear)', '3000万+', '✅', '✅', '✅', '✅', '✅', '✅', '✅', '-', '✅'],
    ['Influencity', '2亿+(TikTok 1.3亿+)', '✅', '✅', '✅', '-', '-', '-', '-', '✅', '-'],
    ['SocialBook', '3亿+', '✅', '✅', '✅', '-', '-', '-', '-', '-', '-'],
    ['FastMoss', '2.5亿达人/5亿商品', '-', '-', '✅', '-', '-', '-', '-', '-', '-'],
    ['EchoTik', '1亿达人/1.8亿商品', '-', '-', '✅', '-', '-', '-', '-', '-', '-'],
    ['Kalodata', '2.5亿达人/2亿商品/4亿视频', '-', '-', '✅', '-', '-', '-', '-', '-', '-'],
    ['Social Blade', '1.6亿创作者', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '-'],
]
for i, row_data in enumerate(platform_data):
    write_row(ws2, i + 4, row_data, bold_first=True)
set_col_widths(ws2, [22, 28, 10, 10, 10, 10, 10, 10, 10, 10, 10])

# ============================================================
# Sheet 3: 网红发现 - 搜索筛选维度
# ============================================================
ws3 = wb.create_sheet("网红发现-搜索筛选维度")
ws3.merge_cells('A1:K1')
ws3['A1'] = '网红发现能力对比 — 搜索筛选维度（✅=已验证存在，-=官网未展示）'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')
ws3.row_dimensions[1].height = 30

headers3 = ['筛选维度', 'NoxInfluencer', 'Modash', 'Heepsy', 'Grin', 'CreatorIQ', 'Aspire', 'Traackr', 'Meltwater', 'Influencity', 'SocialBook']
for col, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=col, value=h)
style_header(ws3, 3, 11)

# Sub-section: 基础筛选
ws3.cell(row=4, column=1, value='【基础筛选维度】')
ws3.cell(row=4, column=1).font = section_font
ws3.merge_cells('A4:K4')
ws3.cell(row=4, column=1).fill = sub_header_fill

filter_data = [
    # 基础筛选
    ['关键词搜索', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '✅'],
    ['话题标签(Hashtag)搜索', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '✅'],
    ['粉丝量范围筛选', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '✅'],
    ['互动率筛选', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '✅'],
    ['地理位置/国家筛选', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '✅'],
    ['语言筛选', '-', '✅', '-', '-', '✅', '-', '-', '✅', '✅', '-'],
    ['兴趣/品类筛选', '✅', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '✅'],
    ['粉丝增长趋势筛选', '✅', '✅', '-', '-', '-', '-', '-', '-', '✅', '-'],
    ['内容类型筛选(Reels/Shorts等)', '-', '✅', '-', '-', '✅', '-', '-', '✅', '-', '-'],
    ['合作品牌历史筛选', '-', '✅', '✅', '-', '✅', '-', '-', '✅', '-', '-'],
]

for i, row_data in enumerate(filter_data):
    write_row(ws3, i + 5, row_data)

# Sub-section: 受众画像筛选
row_offset = 5 + len(filter_data)
ws3.cell(row=row_offset, column=1, value='【受众画像筛选维度】')
ws3.cell(row=row_offset, column=1).font = section_font
ws3.merge_cells(f'A{row_offset}:K{row_offset}')
ws3.cell(row=row_offset, column=1).fill = sub_header_fill

audience_data = [
    ['受众年龄分布', '-', '✅', '-', '-', '✅', '-', '-', '✅', '✅', '-'],
    ['受众性别分布', '-', '✅', '-', '-', '✅', '-', '-', '✅', '✅', '-'],
    ['受众地理位置分布', '-', '✅', '✅', '-', '✅', '-', '-', '✅', '✅', '-'],
    ['受众语言分布', '-', '✅', '-', '-', '✅', '-', '-', '-', '-', '-'],
    ['受众兴趣分类', '-', '✅', '-', '-', '✅', '-', '-', '-', '-', '-'],
    ['受众真实性/假粉率', '-', '✅(AQS+假粉率)', '✅(真实性评分)', '-', '✅(SafeIQ)', '-', '-', '✅(机器人检测)', '-', '✅(IG假粉检测)'],
    ['品牌亲和力分析', '-', '-', '-', '-', '✅', '-', '-', '✅(自定义品牌评分)', '-', '-'],
    ['品牌安全评估', '-', '✅', '✅(品牌安全检查)', '-', '✅(SafeIQ)', '-', '-', '-', '-', '-'],
]

for i, row_data in enumerate(audience_data):
    write_row(ws3, row_offset + 1 + i, row_data)

# Sub-section: AI/高级功能
row_offset2 = row_offset + 1 + len(audience_data)
ws3.cell(row=row_offset2, column=1, value='【AI/高级发现功能】')
ws3.cell(row=row_offset2, column=1).font = section_font
ws3.merge_cells(f'A{row_offset2}:K{row_offset2}')
ws3.cell(row=row_offset2, column=1).fill = sub_header_fill

ai_data = [
    ['AI智能推荐', '-', '✅', '-', '✅(Gia AI)', '✅(3倍速度)', '-', '-', '✅', '✅', '✅(AI产品诊断)'],
    ['相似网红推荐(Lookalike)', '-', '✅', '-', '-', '✅(AI模型)', '-', '-', '-', '✅', '-'],
    ['视觉搜索(上传图片找相似)', '-', '✅(独有)', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['CPM/投放成本估算', '✅(独有)', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['网红排行榜', '✅', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['竞品网红监控', '✅(品牌情报)', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['Influential Fans(已关注品牌的创作者)', '-', '✅(Performance起)', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['Influential Customers(已是客户的创作者)', '-', '✅', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['语义搜索(自然语言)', '-', '-', '-', '-', '-', '-', '-', '✅', '-', '-'],
]

for i, row_data in enumerate(ai_data):
    write_row(ws3, row_offset2 + 1 + i, row_data)

set_col_widths(ws3, [35, 16, 18, 16, 14, 16, 12, 12, 18, 14, 16])

# ============================================================
# Sheet 4: 网红建联 - Email功能详解
# ============================================================
ws4 = wb.create_sheet("网红建联-Email功能详解")
ws4.merge_cells('A1:K1')
ws4['A1'] = '网红建联能力对比 — Email/邮件功能详解'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')
ws4.row_dimensions[1].height = 30

headers4 = ['功能维度', 'NoxInfluencer', 'Modash', 'Heepsy', 'Grin', 'CreatorIQ', 'Aspire', 'Traackr', 'Meltwater', 'Influencity', 'FastMoss']
for col, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=col, value=h)
style_header(ws4, 3, 11)

# Email sending
ws4.cell(row=4, column=1, value='【邮件发送能力】')
ws4.cell(row=4, column=1).font = section_font
ws4.merge_cells('A4:K4')
ws4.cell(row=4, column=1).fill = sub_header_fill

email_send = [
    ['平台内直接发送邮件', '✅ 邮件邀约功能', '✅ 双向同步Gmail/Outlook', '✅ 批量联系+直接消息', '✅ Gmail/Klaviyo集成', '✅', '✅ Contact Hub', '-', '✅', '✅ IRM模块', '✅ MossCreator'],
    ['批量发送邮件', '✅ AI Skills批量', '✅ 一次性发送上百封', '✅ 批量联系', '✅', '-', '-', '-', '-', '-', '✅'],
    ['邮件模板库', '-', '✅ "Save hours with templates"', '-', '✅ 邮件模板', '-', '-', '-', '-', '-', '-'],
    ['个性化变量(名字/粉丝数等)', '✅ AI个性化', '✅ 自动填充', '-', '✅', '-', '-', '-', '-', '-', '-'],
    ['防止重复外展', '-', '✅ 自动检测已有对话', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['收件人无平台标记', '-', '✅ 显示为普通品牌邮箱', '-', '-', '-', '-', '-', '-', '-', '-'],
]

for i, row_data in enumerate(email_send):
    write_row(ws4, i + 5, row_data, bold_first=True)

# Auto follow-up
row_off = 5 + len(email_send)
ws4.cell(row=row_off, column=1, value='【自动跟进序列】')
ws4.cell(row=row_off, column=1).font = section_font
ws4.merge_cells(f'A{row_off}:K{row_off}')
ws4.cell(row=row_off, column=1).fill = sub_header_fill

auto_follow = [
    ['Drip自动跟进序列', '-', '✅ 自动滴灌式跟进，回复后自动停止', '-', '✅ 邮件序列', '-', '✅ Workflow Automation', '-', '-', '✅ 自动邮件追踪', '-'],
    ['自动跟进提醒', '-', '-', '-', '✅', '-', '✅ 智能触发器+提醒', '-', '-', '-', '-'],
    ['AI自动建联', '✅ AI Skills一句指令批量发送', '-', '-', '✅ Gia AI代理', '-', '-', '-', '-', '-', '-'],
    ['AI自动价格谈判', '✅ 行业首创，定义预算底线即可', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
]

for i, row_data in enumerate(auto_follow):
    write_row(ws4, row_off + 1 + i, row_data, bold_first=True)

# WhatsApp & other channels
row_off2 = row_off + 1 + len(auto_follow)
ws4.cell(row=row_off2, column=1, value='【多渠道触达】')
ws4.cell(row=row_off2, column=1).font = section_font
ws4.merge_cells(f'A{row_off2}:K{row_off2}')
ws4.cell(row=row_off2, column=1).fill = sub_header_fill

multi_channel = [
    ['WhatsApp触达', '✅ 独有，解锁WhatsApp功能', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['平台内直接消息', '-', '-', '✅ 直接消息+沟通历史', '-', '-', '-', '-', '-', '-', '-'],
    ['AI开发信生成', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['TikTok Shop邀约插件', '✅ TT Shop邀约插件', '-', '-', '-', '-', '-', '-', '-', '-', '-'],
    ['联系方式获取方式', '消耗credits查看邮箱+解锁WhatsApp', '自动匹配邮箱到350M+资料', '查看联系方式', '自动获取', '自动获取', '自动获取', '-', '自动获取', '自动获取', '自动获取'],
]

for i, row_data in enumerate(multi_channel):
    write_row(ws4, row_off2 + 1 + i, row_data, bold_first=True)

# Creator self-application
row_off3 = row_off2 + 1 + len(multi_channel)
ws4.cell(row=row_off3, column=1, value='【创作者自助申请（Inbound模式）】')
ws4.cell(row=row_off3, column=1).font = section_font
ws4.merge_cells(f'A{row_off3}:K{row_off3}')
ws4.cell(row=row_off3, column=1).fill = sub_header_fill

inbound = [
    ['创作者自助申请门户', '-', '-', '✅ Marketplace+Application Pages', '-', '-', '✅ Inbound Applications+自定义落地页', '-', '-', '-', '-'],
    ['品牌发布活动/创作者申请', '✅ 品牌情报', '-', '✅ Marketplace', '-', '-', '✅', '-', '-', '-', '-'],
    ['自定义落地页', '-', '-', '-', '-', '-', '✅ 可嵌入品牌官网', '-', '-', '-', '-'],
]

for i, row_data in enumerate(inbound):
    write_row(ws4, row_off3 + 1 + i, row_data, bold_first=True)

set_col_widths(ws4, [30, 22, 28, 22, 20, 14, 22, 12, 14, 18, 16])

# ============================================================
# Sheet 5: 网红建联 - CRM与管理
# ============================================================
ws5 = wb.create_sheet("网红建联-CRM与管理")
ws5.merge_cells('A1:K1')
ws5['A1'] = '网红建联能力对比 — CRM、合同、付款、内容管理'
ws5['A1'].font = title_font
ws5['A1'].alignment = Alignment(horizontal='center')
ws5.row_dimensions[1].height = 30

headers5 = ['功能维度', 'NoxInfluencer', 'Modash', 'Heepsy', 'Grin', 'CreatorIQ', 'Aspire', 'Traackr', 'Meltwater', 'Influencity', 'SocialBook']
for col, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=col, value=h)
style_header(ws5, 3, 11)

# CRM
ws5.cell(row=4, column=1, value='【CRM/关系管理】')
ws5.cell(row=4, column=1).font = section_font
ws5.merge_cells('A4:K4')
ws5.cell(row=4, column=1).fill = sub_header_fill

crm_data = [
    ['CRM模块', '✅ 网红收藏夹+推广任务管理', '✅ 基础CRM(管理关系/对话/团队协作/内部备注)', '✅ CRM风格协作工具', '✅ Creator CRM(标签+阶段管理)', '✅ 统一项目管理', '✅ Contact Hub', '-', '✅ 内置CRM', '✅ IRM模块(自定义字段/沟通追踪)', '-'],
    ['团队协作/子账号', '✅ 子账号功能', '✅ 共享对话/备注/交接', '-', '✅', '✅', '✅', '-', '-', '-', '-'],
    ['沟通历史记录', '-', '✅', '✅ 沟通历史追踪', '✅', '✅', '✅', '-', '✅', '✅ 自动邮件追踪监控', '-'],
    ['网红分组/标签管理', '✅ 收藏分组', '✅', '-', '✅ 标签和阶段', '✅', '✅', '-', '✅', '✅ 自定义字段', '-'],
]

for i, row_data in enumerate(crm_data):
    write_row(ws5, i + 5, row_data, bold_first=True)

# Contract & Payment
row_off = 5 + len(crm_data)
ws5.cell(row=row_off, column=1, value='【合同与付款管理】')
ws5.cell(row=row_off, column=1).font = section_font
ws5.merge_cells(f'A{row_off}:K{row_off}')
ws5.cell(row=row_off, column=1).fill = sub_header_fill

payment_data = [
    ['合同管理', '-', '-', '✅ 合同管理', '✅ DocuSign集成', '-', '✅ Briefs & Agreements', '-', '✅', '-', '-'],
    ['付款管理', '✅ Payment功能+推广订单', '✅ 180+国家全球付款/合并发票', '✅ Payments模块(无内置处理器)', '✅ Creator Payments(PayPal+1099税务)', '✅ CreatorIQ Pay', '✅', '-', '✅ Tipalti+Gigapay', '✅ 多币种批量付款', '-'],
    ['联盟营销/佣金追踪', '✅ Shopify联盟营销', '✅ 联盟管理(0%手续费)', '-', '✅ 联盟链接/折扣码/佣金/深度链接/高级归因', '-', '✅ 联盟营销', '-', '-', '-', '-'],
]

for i, row_data in enumerate(payment_data):
    write_row(ws5, row_off + 1 + i, row_data, bold_first=True)

# Content management
row_off2 = row_off + 1 + len(payment_data)
ws5.cell(row=row_off2, column=1, value='【内容管理】')
ws5.cell(row=row_off2, column=1).font = section_font
ws5.merge_cells(f'A{row_off2}:K{row_off2}')
ws5.cell(row=row_off2, column=1).fill = sub_header_fill

content_data = [
    ['UGC内容收集与管理', '-', '-', '-', '✅ 内容收集/集中存储/跨渠道复用/使用权利管理', '-', '-', '-', '-', '-', '-'],
    ['内容审批工作流', '-', '-', '✅ 内容审批', '✅ Growth方案起', '-', '-', '-', '✅', '-', '-'],
    ['内容自动追踪', '✅ 追踪网红内容表现', '✅ 自动收集创作者发布内容', '✅ Tracking Content', '✅', '-', '-', '-', '✅', '-', '-'],
    ['产品寄送管理', '-', '✅ Shopify产品赠送', '-', '✅ 产品寄送自动化', '-', '✅ 自动发货物流追踪', '-', '✅ 产品赠送', '-'],
    ['社交监听', '-', '-', '-', '✅ Social Listening(hashtag限制)', '-', '-', '-', '✅ Meltwater大平台', '✅ 1亿+来源', '-'],
]

for i, row_data in enumerate(content_data):
    write_row(ws5, row_off2 + 1 + i, row_data, bold_first=True)

# E-commerce
row_off3 = row_off2 + 1 + len(content_data)
ws5.cell(row=row_off3, column=1, value='【电商集成详情】')
ws5.cell(row=row_off3, column=1).font = section_font
ws5.merge_cells(f'A{row_off3}:K{row_off3}')
ws5.cell(row=row_off3, column=1).fill = sub_header_fill

ecom_data = [
    ['Shopify集成', '✅ 联盟营销', '✅ 产品赠送/销售追踪/折扣码/联盟管理', '✅ Shopify Detection', '✅ 深度集成(CreatorCommerce联盟着陆页)', '-', '✅ 深度集成(一键连接/买家数据/自动发货/批量优惠码/全链路追踪)', '-', '✅', '✅ 电商店铺集成(Seeding)', '-'],
    ['WooCommerce集成', '-', '-', '-', '-', '-', '✅', '-', '✅', '-', '-'],
    ['Amazon集成', '-', '-', '✅ Amazon销售归因', '-', '-', '-', '-', '-', '-', '-'],
    ['TikTok Shop集成', '✅ TT Shop邀约插件', '-', '✅ TikTok Shop归因', '-', '-', '-', '-', '-', '-', '-'],
    ['销售归因追踪', '-', '✅', '✅ Shopify/Amazon/TikTok Shop三平台', '✅ 高级归因', '-', '✅ 全链路', '-', '-', '-', '-'],
]

for i, row_data in enumerate(ecom_data):
    write_row(ws5, row_off3 + 1 + i, row_data, bold_first=True)

set_col_widths(ws5, [28, 22, 28, 28, 28, 18, 28, 12, 18, 20, 12])

# ============================================================
# Sheet 6: 定价对比
# ============================================================
ws6 = wb.create_sheet("定价对比")
ws6.merge_cells('A1:F1')
ws6['A1'] = '定价对比（已验证数据）'
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

headers6 = ['平台', '免费/试用', '入门价', '中端价', '高端价', '计费方式']
for col, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=col, value=h)
style_header(ws6, 3, 6)

price_data = [
    ['Modash', '14天试用(20资料/6邮箱/10追踪)', '$199/月(年付) Essentials', '$499/月(年付) Performance', '$14,700+/年 Enterprise', '月付/年付'],
    ['Heepsy', '免费试用', '$69/月', '-', '-', '-'],
    ['Grin', '30天试用(15创作者/150CRM)', '$399/月 Lite(50创作者)', '$1,149/月 Growth(200创作者)', '$1,799/月 Complete(400创作者)', '按月随时取消'],
    ['Grin联盟计划', '永久免费(0%分成/无合同)', '$0', '-', '-', '仅Shopify'],
    ['EchoTik', '免费版(基础筛选)', '$9.9/月 Basic(100条导出/天)', '$19.1/月 Pro(2000条/天)', '$29.1/月 Enterprise(无限)', '月付/7天退款'],
    ['CreatorIQ', '-', '不公开(联系销售)', '-', '-', '企业级'],
    ['Aspire', '-', '不公开(联系销售)', '-', '-', '-'],
    ['Traackr', '-', '不公开(联系销售)', '-', '-', '企业级'],
    ['Meltwater', '-', '不公开(联系销售)', '-', '-', '企业级'],
    ['Influencity', '7天免费试用', '不公开(需JS加载)', '不公开', '不公开', '无强制合约/随时取消'],
]

for i, row_data in enumerate(price_data):
    write_row(ws6, i + 4, row_data, bold_first=True)
set_col_widths(ws6, [20, 32, 28, 28, 28, 22])

# ============================================================
# Sheet 7: 差异化功能
# ============================================================
ws7 = wb.create_sheet("差异化功能")
ws7.merge_cells('A1:C1')
ws7['A1'] = '独有/特色功能对比'
ws7['A1'].font = title_font
ws7['A1'].alignment = Alignment(horizontal='center')

headers7 = ['平台', '独有/特色功能', '竞品是否有']
for col, h in enumerate(headers7, 1):
    ws7.cell(row=3, column=col, value=h)
style_header(ws7, 3, 3)

diff_data = [
    ['NoxInfluencer', 'WhatsApp触达', '❌ 无竞品有'],
    ['NoxInfluencer', 'AI自动价格谈判(Skills)', '❌ 行业首创'],
    ['NoxInfluencer', 'CPM估算', '❌ 无竞品有'],
    ['NoxInfluencer', 'TT Shop邀约插件', '部分(TikTok工具类似)'],
    ['Modash', '视觉搜索(上传图片找相似创作者)', '❌ 无竞品有'],
    ['Modash', 'Influential Fans/Customers', '部分(Aspire有类似)'],
    ['Modash', 'Gmail/Outlook双向同步(非集成)', '❌ 独有实现方式'],
    ['Heepsy', 'Marketplace(创作者主动申请)', '类似Aspire的Inbound'],
    ['Heepsy', 'Shopify+Amazon+TikTok Shop三平台归因', '❌ 独有组合'],
    ['Grin', 'AI代理Gia(4层级智能助手)', 'Nox有AI Skills，其他无'],
    ['Grin', '永久免费联盟计划(0%分成)', '❌ 无竞品有'],
    ['Grin', 'CreatorCommerce联合品牌联盟着陆页', '❌ 独有'],
    ['CreatorIQ', 'SafeIQ品牌安全(AI驱动)', '部分平台有类似功能'],
    ['CreatorIQ', 'The Creator Graph(内容优先智能基础设施)', '❌ 独有架构'],
    ['Aspire', 'Inbound Applications自定义落地页', 'Heepsy有Marketplace'],
    ['Aspire', 'Contact Hub统一管理所有创作者关系', '部分CRM类似'],
    ['Meltwater', 'True Reach算法(真实触达评估)', '❌ 独有'],
    ['Meltwater', 'Influencer Score(0-100分AI驱动)', '❌ 独有'],
    ['Meltwater', '自定义品牌评分(受众与品牌匹配度)', '❌ 独有'],
    ['Meltwater', '媒体情报生态(社交聆听+新闻分析+网红管理)', '❌ 独有'],
    ['Influencity', '社交聆听(1亿+来源)集成网红营销', 'Meltwater有类似'],
    ['FastMoss', 'MossCreator达人批量邮件(独立产品)', 'Modash有类似'],
    ['FastMoss', 'TikTok视频无水印下载', '❌ 独有'],
    ['EchoTik', '全套AI工具(仿写/脚本/标题/开发信/图片翻译/擦除)', '部分(FastMoss有AI)'],
    ['EchoTik', '国家大盘数据(如美国周GMV $2.429亿)', '❌ 独有'],
    ['Kalodata', 'TikTok Shop早期核心团队背景', '❌ 独有背景'],
    ['SocialBook', 'AI产品诊断(分析竞品+推荐红人+生成策略)', '❌ 独有'],
    ['SocialBook', '全托管红人营销服务(Managed Campaign)', '部分有托管服务'],
    ['SocialBook', '红人价格评估工具(帮助红人评估帖子定价)', '❌ 独有'],
]

for i, row_data in enumerate(diff_data):
    write_row(ws7, i + 4, row_data, bold_first=True)
set_col_widths(ws7, [20, 48, 28])

# ============================================================
# Sheet 8: 场景推荐
# ============================================================
ws8 = wb.create_sheet("场景推荐")
ws8.merge_cells('A1:D1')
ws8['A1'] = '按使用场景推荐选型'
ws8['A1'].font = title_font
ws8['A1'].alignment = Alignment(horizontal='center')

headers8 = ['使用场景', '推荐平台', '推荐理由', '关键功能']
for col, h in enumerate(headers8, 1):
    ws8.cell(row=3, column=col, value=h)
style_header(ws8, 3, 4)

scenario_data = [
    ['中国出海品牌（预算有限）', 'NoxInfluencer + EchoTik', '中文友好、WhatsApp双渠道、$9.9/月起', '邮件邀约+WhatsApp+AI Skills+TT Shop插件'],
    ['TikTok Shop电商卖家', 'EchoTik / FastMoss / Kalodata', 'TikTok深度数据、选品+达人+直播全链路', '商品分析+达人带货数据+直播分析+AI工具'],
    ['中小企业（全球网红营销）', 'Heepsy → Modash', '$69起步→$199规模化', '假粉检测+Shopify归因+Marketplace+Drip序列'],
    ['DTC电商品牌（Shopify为主）', 'Grin / Aspire', '免费联盟计划/AI代理/Inbound模式', 'Shopify深度集成+产品寄送+UGC管理+联盟营销'],
    ['企业级/跨国品牌', 'CreatorIQ / Meltwater', 'AI驱动/True Reach/媒体情报/190+国家', 'SafeIQ品牌安全+CreatorIQ Pay+社交聆听+竞品基准'],
    ['需要反欺诈/数据真实性', 'Modash / Heepsy', 'AQS假粉率检测/真实性评分/品牌安全', '受众质量分析+假粉百分比+品牌安全检查'],
    ['免费查看社交媒体数据', 'Social Blade', '1.6亿创作者、多平台、免费', '统计追踪+排行榜+收入预估+Chrome扩展'],
]

for i, row_data in enumerate(scenario_data):
    write_row(ws8, i + 4, row_data, bold_first=True)
set_col_widths(ws8, [28, 28, 35, 40])

# Save
excel_path = r'D:\program\ai-novel\influencer_platform_report.xlsx'
wb.save(excel_path)
print(f"Excel saved: {excel_path}")
print("Done!")
